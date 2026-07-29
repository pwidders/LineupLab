import json
from datetime import datetime
from io import BytesIO

import altair as alt
import pandas as pd
import streamlit as st

from contest_history_store import load_contest_history


def calculate_overall_metrics(history: pd.DataFrame) -> dict:
    """
    Calculate portfolio-level DFS performance metrics.

    Keeping calculations separate from Streamlit rendering allows
    this function to be reused later for filters, exports, and tests.
    """
    total_contests = len(history)

    total_fees = float(history["entry_fee"].sum())
    total_winnings = float(history["winnings"].sum())
    net_profit = total_winnings - total_fees

    roi = (
        net_profit / total_fees
        if total_fees > 0
        else 0.0
    )

    cash_rate = (
        float((history["winnings"] > 0).mean())
        if total_contests > 0
        else 0.0
    )

    average_points = (
        float(history["points"].mean())
        if "points" in history.columns
        else 0.0
    )

    # rank_percentile is stored as rank / field_size.
    # Subtracting from 1 converts it into a finish score:
    # 100% = best possible finish, 0% = last place.
    average_finish_percentile = (
        float((1 - history["rank_percentile"]).mean())
        if "rank_percentile" in history.columns
        else 0.0
    )

    return {
        "total_fees": total_fees,
        "total_winnings": total_winnings,
        "net_profit": net_profit,
        "roi": roi,
        "cash_rate": cash_rate,
        "total_contests": total_contests,
        "average_points": average_points,
        "average_finish_percentile": (
            average_finish_percentile
        ),
    }

def build_slate_history(history: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate contest history into one row per slate.

    This becomes the foundation for all drill-down analytics.
    """

    if history.empty:
        return pd.DataFrame()

    grouped = (
        history.groupby("slate_date", as_index=False)
        .agg(
            contests=("entry_name", "count"),
            total_fees=("entry_fee", "sum"),
            total_winnings=("winnings", "sum"),
            average_points=("points", "mean"),
        )
    )

    grouped["profit"] = (
        grouped["total_winnings"] - grouped["total_fees"]
    )

    grouped["roi"] = grouped.apply(
        lambda row:
            row["profit"] / row["total_fees"]
            if row["total_fees"] > 0
            else 0,
        axis=1,
    )

    grouped = grouped.sort_values(
        "slate_date",
        ascending=False,
    )

    return grouped


def build_player_history(history: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten player_results JSON from contest history into
    one row per player appearance.
    """

    if history.empty or "player_results" not in history.columns:
        return pd.DataFrame()

    player_rows = []

    for _, contest in history.iterrows():
        results = contest.get("player_results")

        if not isinstance(results, list):
            continue

        for player in results:
            if not isinstance(player, dict):
                continue

            player_rows.append(
                {
                    "player": player.get("player", ""),
                    "roster_position": player.get(
                        "roster_position",
                        "",
                    ),
                    "team": (
                        player.get("team")
                        or player.get("team_abbrev")
                        or player.get("team_abbreviation")
                        or player.get("team_name")
                        or ""
                    ),
                    "salary": player.get("salary", 0),
                    "projection": (
                        player.get("projection")
                        or player.get("projected_points")
                        or 0
                    ),
                    "ownership": player.get("ownership", 0),
                    "fpts": player.get("fpts", 0),
                    "slate_date": contest.get("slate_date"),
                    "contest_type": contest.get("contest_type"),
                    "entry_name": contest.get("entry_name", ""),
                    "entry_fee": contest.get("entry_fee", 0),
                    "winnings": contest.get("winnings", 0),
                    "profit": contest.get("profit", 0),
                    "lineup_id": contest.get("lineup_id"),
                }
            )

    if not player_rows:
        return pd.DataFrame()

    players = pd.DataFrame(player_rows)

    players["ownership"] = pd.to_numeric(
        players["ownership"],
        errors="coerce",
    ).fillna(0)

    players["salary"] = pd.to_numeric(
        players["salary"],
        errors="coerce",
    ).fillna(0)

    players["projection"] = pd.to_numeric(
        players["projection"],
        errors="coerce",
    ).fillna(0)

    players["fpts"] = pd.to_numeric(
        players["fpts"],
        errors="coerce",
    ).fillna(0)

    players["player_key"] = (
        players["player"]
        .astype(str)
        .str.strip()
        .str.lower()
    )

    players = players.drop_duplicates(
        subset=[
            "slate_date",
            "contest_type",
            "entry_name",
            "lineup_id",
            "player_key",
        ],
        keep="first",
    )

    players = players.drop(columns=["player_key"])

    return players


def apply_dashboard_filters(history: pd.DataFrame) -> pd.DataFrame:
    """
    Render the shared Performance Dashboard filters and return
    the filtered contest-history rows used by every section.
    """

    filtered = history.copy()

    filtered["slate_date"] = pd.to_datetime(
        filtered["slate_date"],
        errors="coerce",
    )

    filtered = filtered.dropna(subset=["slate_date"])

    minimum_date = filtered["slate_date"].min().date()
    maximum_date = filtered["slate_date"].max().date()

    contest_types = sorted(
        filtered["contest_type"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    slate_options = sorted(
        filtered["slate_date"]
        .dt.date
        .unique()
        .tolist(),
        reverse=True,
    )

    st.markdown("### Filters")

    filter_columns = st.columns([1.35, 1.35, 1.35, 1])

    selected_dates = filter_columns[0].date_input(
        "Date Range",
        value=(minimum_date, maximum_date),
        min_value=minimum_date,
        max_value=maximum_date,
        key="performance_date_range",
    )

    selected_contests = filter_columns[1].multiselect(
        "Contest Type",
        options=contest_types,
        default=contest_types,
        key="performance_contest_types",
    )

    selected_slates = filter_columns[2].multiselect(
        "Slate",
        options=slate_options,
        default=[],
        format_func=lambda value: value.strftime("%Y-%m-%d"),
        placeholder="All slates",
        key="performance_slates",
    )

    result_filter = filter_columns[3].selectbox(
        "Result",
        options=["All", "Cashed", "Did Not Cash"],
        key="performance_result_filter",
    )

    if isinstance(selected_dates, (tuple, list)):
        if len(selected_dates) == 2:
            start_date, end_date = selected_dates
        elif len(selected_dates) == 1:
            start_date = selected_dates[0]
            end_date = selected_dates[0]
        else:
            start_date = minimum_date
            end_date = maximum_date
    else:
        start_date = selected_dates
        end_date = selected_dates

    filtered = filtered[
        filtered["slate_date"].dt.date.between(
            start_date,
            end_date,
        )
    ]

    if selected_contests:
        filtered = filtered[
            filtered["contest_type"]
            .astype(str)
            .isin(selected_contests)
        ]
    else:
        filtered = filtered.iloc[0:0]

    if selected_slates:
        filtered = filtered[
            filtered["slate_date"].dt.date.isin(selected_slates)
        ]

    if result_filter == "Cashed":
        filtered = filtered[filtered["winnings"] > 0]
    elif result_filter == "Did Not Cash":
        filtered = filtered[filtered["winnings"] <= 0]

    return filtered.copy()


def build_stack_export(players: pd.DataFrame) -> pd.DataFrame:
    """Build one export row per contest entry's primary hitter stack."""
    if players.empty:
        return pd.DataFrame()

    stack_players = players.copy()
    stack_players["team"] = (
        stack_players["team"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )
    stack_players["normalized_position"] = (
        stack_players["roster_position"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )
    stack_players = stack_players[
        ~stack_players["normalized_position"].isin({"P", "SP", "RP"})
    ].copy()
    stack_players = stack_players[stack_players["team"] != ""].copy()

    if stack_players.empty:
        return pd.DataFrame()

    group_columns = [
        "slate_date",
        "contest_type",
        "entry_name",
        "lineup_id",
        "lineup_key",
        "entry_fee",
        "winnings",
        "profit",
        "cashed",
    ]

    team_results = (
        stack_players.groupby(
            group_columns + ["team"],
            dropna=False,
            as_index=False,
        )
        .agg(
            hitter_count=("player", "nunique"),
            stack_dk_points=("fpts", "sum"),
        )
        .sort_values(
            [
                "lineup_key",
                "contest_type",
                "entry_name",
                "hitter_count",
                "stack_dk_points",
                "team",
            ],
            ascending=[True, True, True, False, False, True],
        )
    )

    primary = (
        team_results.groupby(
            ["lineup_key", "contest_type", "entry_name"],
            dropna=False,
            as_index=False,
        )
        .first()
    )

    lineup_points = (
        players.groupby(
            ["lineup_key", "contest_type", "entry_name"],
            dropna=False,
            as_index=False,
        )
        .agg(lineup_dk_points=("fpts", "sum"))
    )

    primary = primary.merge(
        lineup_points,
        on=["lineup_key", "contest_type", "entry_name"],
        how="left",
    )
    primary["primary_stack"] = (
        primary["team"]
        + " "
        + primary["hitter_count"].astype(int).astype(str)
        + "-man"
    )
    primary["roi"] = primary.apply(
        lambda row: (
            row["profit"] / row["entry_fee"]
            if row["entry_fee"] > 0
            else 0.0
        ),
        axis=1,
    )

    export_columns = [
        "slate_date",
        "contest_type",
        "entry_name",
        "lineup_id",
        "team",
        "hitter_count",
        "primary_stack",
        "stack_dk_points",
        "lineup_dk_points",
        "entry_fee",
        "winnings",
        "profit",
        "roi",
        "cashed",
    ]
    return primary[export_columns].sort_values(
        ["slate_date", "contest_type", "entry_name"],
        ascending=[False, True, True],
    )


def _excel_safe_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    """Convert nested JSON values and timezone-aware dates for Excel."""
    safe = frame.copy()

    for column in safe.columns:
        if safe[column].dtype == "object":
            safe[column] = safe[column].apply(
                lambda value: json.dumps(value, default=str)
                if isinstance(value, (dict, list, tuple, set))
                else value
            )

        if pd.api.types.is_datetime64_any_dtype(safe[column]):
            try:
                safe[column] = safe[column].dt.tz_localize(None)
            except (TypeError, AttributeError):
                pass

    return safe


def _style_excel_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    """Apply readable formatting to one openpyxl worksheet."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for index, column in enumerate(dataframe.columns, start=1):
        values = [str(column)]
        values.extend(
            "" if pd.isna(value) else str(value)
            for value in dataframe[column].head(250)
        )
        width = min(max(len(value) for value in values) + 2, 38)
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            width,
            10,
        )

        name = str(column).lower()
        if any(token in name for token in ["fee", "winning", "profit", "salary"]):
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif any(token in name for token in ["roi", "cash_rate", "percentile"]):
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = "0.0%"
        elif any(token in name for token in ["points", "projection", "ownership"]):
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = "0.00"


def generate_algorithm_review_workbook(
    filtered_history: pd.DataFrame,
    metrics: dict,
    players: pd.DataFrame,
) -> bytes:
    """Create the filtered Algorithm Review Report as an Excel workbook."""
    generated_at = datetime.now().astimezone()

    contest_export = filtered_history.copy()
    contest_export["roi"] = contest_export.apply(
        lambda row: (
            row["profit"] / row["entry_fee"]
            if row.get("entry_fee", 0) > 0
            else 0.0
        ),
        axis=1,
    )
    contest_export["finish_percentile"] = (
        1 - pd.to_numeric(
            contest_export.get("rank_percentile", 0),
            errors="coerce",
        ).fillna(0)
    )
    contest_export["cashed"] = contest_export["winnings"] > 0

    player_export = players.copy()
    if not player_export.empty:
        player_export["projection_difference"] = (
            player_export["fpts"] - player_export["projection"]
        )
        player_export["value_per_1000"] = player_export.apply(
            lambda row: (
                row["fpts"] / (row["salary"] / 1000)
                if row["salary"] > 0
                else 0.0
            ),
            axis=1,
        )
        player_export = player_export.sort_values(
            ["slate_date", "contest_type", "entry_name", "roster_position"],
            ascending=[False, True, True, True],
        )

    pitcher_positions = {"P", "SP", "RP"}
    pitcher_export = player_export[
        player_export["roster_position"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(pitcher_positions)
    ].copy() if not player_export.empty else pd.DataFrame()

    stack_export = build_stack_export(players)

    unique_lineups = 0
    if not players.empty and "lineup_key" in players.columns:
        unique_lineups = int(players["lineup_key"].nunique())

    date_min = pd.to_datetime(filtered_history["slate_date"]).min()
    date_max = pd.to_datetime(filtered_history["slate_date"]).max()
    contest_types = ", ".join(
        sorted(
            filtered_history["contest_type"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )
    )

    summary = pd.DataFrame(
        [
            ("Report Generated", generated_at.strftime("%Y-%m-%d %I:%M %p %Z")),
            ("Date Range", f"{date_min:%Y-%m-%d} through {date_max:%Y-%m-%d}"),
            ("Contest Types", contest_types or "All"),
            ("Entries", metrics["total_contests"]),
            ("Unique Lineups", unique_lineups),
            ("Total Fees", metrics["total_fees"]),
            ("Total Winnings", metrics["total_winnings"]),
            ("Net Profit", metrics["net_profit"]),
            ("ROI", metrics["roi"]),
            ("Cash Rate", metrics["cash_rate"]),
            ("Average DK Points", metrics["average_points"]),
            ("Average Finish Percentile", metrics["average_finish_percentile"]),
        ],
        columns=["Metric", "Value"],
    )

    version = pd.DataFrame(
        [
            ("LineupLab Report Version", "2.1"),
            ("Performance Dashboard Version", "2.1"),
            ("Generated On", generated_at.isoformat()),
            ("Contest Rows", len(contest_export)),
            ("Player Rows", len(player_export)),
            ("Pitcher Rows", len(pitcher_export)),
            ("Stack Rows", len(stack_export)),
        ],
        columns=["Field", "Value"],
    )

    raw_history = _excel_safe_dataframe(filtered_history)
    contest_export = _excel_safe_dataframe(contest_export)
    player_export = _excel_safe_dataframe(player_export)
    pitcher_export = _excel_safe_dataframe(pitcher_export)
    stack_export = _excel_safe_dataframe(stack_export)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets = {
            "Summary": summary,
            "Contest Results": contest_export,
            "Player Results": player_export,
            "Pitcher Results": pitcher_export,
            "Stack Results": stack_export,
            "Raw History": raw_history,
            "Version": version,
        }

        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_excel_sheet(writer.sheets[sheet_name], dataframe)

        summary_sheet = writer.sheets["Summary"]
        for row in range(2, summary_sheet.max_row + 1):
            metric_name = summary_sheet.cell(row=row, column=1).value
            value_cell = summary_sheet.cell(row=row, column=2)
            if metric_name in {"Total Fees", "Total Winnings", "Net Profit"}:
                value_cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
            elif metric_name in {"ROI", "Cash Rate", "Average Finish Percentile"}:
                value_cell.number_format = "0.0%"
            elif metric_name == "Average DK Points":
                value_cell.number_format = "0.00"

    output.seek(0)
    return output.getvalue()


def format_profit(value: float) -> str:
    if value > 0:
        return f"+${value:,.2f}"

    if value < 0:
        return f"-${abs(value):,.2f}"

    return "$0.00"


def render_performance_center():
    st.header("📊 Performance Dashboard")

    try:
        history = load_contest_history()
    except Exception as exc:
        st.error(
            "Performance Center could not load your "
            f"contest history: {exc}"
        )
        return

    if history.empty:
        st.info(
            "No contest history has been saved yet. "
            "Upload a DraftKings contest export in "
            "Contest Review and save it to Contest History."
        )
        return

    filtered_history = apply_dashboard_filters(history)

    st.caption(
        f"Showing {len(filtered_history):,} of "
        f"{len(history):,} contest entries."
    )

    if filtered_history.empty:
        st.warning(
            "No contest entries match the selected filters."
        )
        return

    metrics = calculate_overall_metrics(filtered_history)

    report_players = build_player_history(filtered_history)

    if not report_players.empty:
        report_players = report_players.copy()
        report_players["entry_fee"] = pd.to_numeric(
            report_players["entry_fee"],
            errors="coerce",
        ).fillna(0)
        report_players["winnings"] = pd.to_numeric(
            report_players["winnings"],
            errors="coerce",
        ).fillna(0)
        report_players["profit"] = (
            report_players["winnings"]
            - report_players["entry_fee"]
        )
        report_players["cashed"] = (
            report_players["winnings"] > 0
        )
        report_players["lineup_key"] = (
            report_players["slate_date"].astype(str)
            + "|"
            + report_players["lineup_id"].fillna("").astype(str)
        )
        missing_report_lineup_id = (
            report_players["lineup_id"].isna()
            | (
                report_players["lineup_id"]
                .astype(str)
                .str.strip()
                == ""
            )
        )
        report_players.loc[
            missing_report_lineup_id,
            "lineup_key",
        ] = (
            report_players.loc[
                missing_report_lineup_id,
                "slate_date",
            ].astype(str)
            + "|"
            + report_players.loc[
                missing_report_lineup_id,
                "entry_name",
            ].astype(str)
        )

    st.markdown("### 📥 Algorithm Review Report")
    st.caption(
        "Download the currently filtered contest, player, pitcher, "
        "stack, and raw history data for optimizer analysis."
    )

    try:
        report_bytes = generate_algorithm_review_workbook(
            filtered_history=filtered_history,
            metrics=metrics,
            players=report_players,
        )
        report_timestamp = datetime.now().strftime(
            "%Y-%m-%d_%H%M"
        )
        st.download_button(
            label="📥 Generate Algorithm Review Report",
            data=report_bytes,
            file_name=(
                "Algorithm_Review_Report_"
                f"{report_timestamp}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
        )
    except Exception as exc:
        st.error(f"Could not generate the report: {exc}")

    st.subheader("At a Glance")

    cards = st.columns(5)

    cards[0].metric(
        "Lifetime Profit",
        format_profit(metrics["net_profit"]),
    )

    cards[1].metric(
        "ROI",
        f"{metrics['roi']:.1%}",
    )

    cards[2].metric(
        "Cash Rate",
        f"{metrics['cash_rate']:.1%}",
    )

    cards[3].metric(
        "Entries",
        f"{metrics['total_contests']:,}",
    )

    cards[4].metric(
        "Average DK Points",
        f"{metrics['average_points']:.2f}",
    )

    st.divider()

    st.subheader("📈 Bankroll Curve")

    slates = build_slate_history(filtered_history)

    if slates.empty:
        st.info("No slates found.")
    else:
        profit_curve = (
            slates[
                ["slate_date", "profit"]
            ]
            .copy()
            .sort_values("slate_date")
        )

        profit_curve["slate_date"] = pd.to_datetime(
            profit_curve["slate_date"],
            errors="coerce",
        )

        profit_curve = profit_curve.dropna(
            subset=["slate_date"]
        )

        profit_curve["cumulative_profit"] = (
            profit_curve["profit"].cumsum()
        )

        bankroll_chart = (
            alt.Chart(profit_curve)
            .mark_line(point=True)
            .encode(
                x=alt.X(
                    "slate_date:T",
                    title="Slate Date",
                    axis=alt.Axis(
                        format="%b %d",
                        labelAngle=0,
                    ),
                ),
                y=alt.Y(
                    "cumulative_profit:Q",
                    title="Profit ($)",
                ),
                tooltip=[
                    alt.Tooltip(
                        "slate_date:T",
                        title="Slate",
                        format="%b %d, %Y",
                    ),
                    alt.Tooltip(
                        "profit:Q",
                        title="Slate Profit",
                        format="$.2f",
                    ),
                    alt.Tooltip(
                        "cumulative_profit:Q",
                        title="Running Profit",
                        format="$.2f",
                    ),
                ],
            )
            .properties(height=280)
        )

        st.altair_chart(
            bankroll_chart,
            use_container_width=True,
        )

        st.caption(
            "Each point reflects your running lifetime profit "
            "after all contests from that slate are included."
        )

    st.markdown("### 🏟 Contest Performance")

    contest_performance = filtered_history.copy()

    contest_performance["roi"] = contest_performance.apply(
        lambda row: (
            row["profit"] / row["entry_fee"]
            if row["entry_fee"] > 0
            else 0.0
        ),
        axis=1,
    )

    contest_performance["finish"] = (
        contest_performance["rank"].astype(int).astype(str)
        + " / "
        + contest_performance["field_size"].astype(int).astype(str)
    )

    contest_performance["cash_result"] = (
        contest_performance["winnings"] > 0
    ).map(
        {
            True: "✅",
            False: "❌",
        }
    )

    contest_performance = contest_performance.sort_values(
        ["slate_date", "contest_type"],
        ascending=[False, True],
    )

    contest_display = contest_performance[
        [
            "slate_date",
            "contest_type",
            "entry_name",
            "entry_fee",
            "winnings",
            "profit",
            "roi",
            "points",
            "finish",
            "cash_result",
        ]
    ].copy()

    contest_display["slate_date"] = pd.to_datetime(
        contest_display["slate_date"]
    ).dt.strftime("%Y-%m-%d")

    contest_display["entry_fee"] = contest_display["entry_fee"].map(
        lambda value: f"${value:.2f}"
    )

    contest_display["winnings"] = contest_display["winnings"].map(
        lambda value: f"${value:.2f}"
    )

    contest_display["profit"] = contest_display["profit"].map(
        format_profit
    )

    contest_display["roi"] = contest_display["roi"].map(
        lambda value: f"{value:.1%}"
    )

    contest_display["points"] = contest_display["points"].map(
        lambda value: f"{value:.2f}"
    )

    contest_display = contest_display.rename(
        columns={
            "slate_date": "Slate",
            "contest_type": "Contest",
            "entry_name": "Entry",
            "entry_fee": "Fee",
            "winnings": "Won",
            "profit": "Profit",
            "roi": "ROI",
            "points": "DK Points",
            "finish": "Finish",
            "cash_result": "Cash",
        }
    )

    st.dataframe(
        contest_display,
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    st.subheader("👤 Player Performance")

    players = build_player_history(filtered_history)

    if players.empty:
        st.info(
            "Player performance will appear after contests "
            "with player-results data are saved."
        )
        return

    players = players.copy()

    players["entry_fee"] = pd.to_numeric(
        players["entry_fee"],
        errors="coerce",
    ).fillna(0)

    players["winnings"] = pd.to_numeric(
        players["winnings"],
        errors="coerce",
    ).fillna(0)

    players["profit"] = (
        players["winnings"] - players["entry_fee"]
    )

    players["cashed"] = players["winnings"] > 0

    players["lineup_key"] = (
        players["slate_date"].astype(str)
        + "|"
        + players["lineup_id"].fillna("").astype(str)
    )

    missing_lineup_id = (
        players["lineup_id"].isna()
        | (
            players["lineup_id"]
            .astype(str)
            .str.strip()
            == ""
        )
    )

    players.loc[missing_lineup_id, "lineup_key"] = (
        players.loc[missing_lineup_id, "slate_date"].astype(str)
        + "|"
        + players.loc[missing_lineup_id, "entry_name"].astype(str)
    )

    player_summary = (
        players.groupby("player", as_index=False)
        .agg(
            uses=("lineup_key", "nunique"),
            entries=("player", "size"),
            average_fpts=("fpts", "mean"),
            average_ownership=("ownership", "mean"),
            total_fees=("entry_fee", "sum"),
            profit=("profit", "sum"),
            cash_rate=("cashed", "mean"),
        )
    )

    player_summary = player_summary[
        player_summary["player"].astype(str).str.strip() != ""
    ].copy()

    player_summary["roi"] = player_summary.apply(
        lambda row: (
            row["profit"] / row["total_fees"]
            if row["total_fees"] > 0
            else 0.0
        ),
        axis=1,
    )

    player_summary = player_summary.sort_values(
        ["uses", "average_fpts"],
        ascending=[False, False],
    )

    player_display = player_summary[
        [
            "player",
            "uses",
            "entries",
            "average_fpts",
            "average_ownership",
            "profit",
            "roi",
            "cash_rate",
        ]
    ].copy()

    player_display["average_fpts"] = (
        player_display["average_fpts"].round(2)
    )

    player_display["average_ownership"] = (
        player_display["average_ownership"].map(
            lambda value: f"{value:.1f}%"
        )
    )

    player_display["profit"] = player_display["profit"].map(
        format_profit
    )

    player_display["roi"] = player_display["roi"].map(
        lambda value: f"{value:.1%}"
    )

    player_display["cash_rate"] = player_display["cash_rate"].map(
        lambda value: f"{value:.1%}"
    )

    player_display = player_display.rename(
        columns={
            "player": "Player",
            "uses": "Uses",
            "entries": "Entries",
            "average_fpts": "Avg DK Points",
            "average_ownership": "Avg Ownership",
            "profit": "Profit",
            "roi": "ROI",
            "cash_rate": "Cash Rate",
        }
    )

    st.caption(
        "Uses counts unique lineups. Entries counts every contest "
        "entry containing that player."
    )

    st.dataframe(
        player_display,
        use_container_width=True,
        hide_index=True,
        height=420,
    )


    st.divider()

    st.subheader("⚾ Pitcher Performance")

    pitcher_positions = {"P", "SP", "RP"}

    pitchers = players[
        players["roster_position"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(pitcher_positions)
    ].copy()

    if pitchers.empty:
        st.info(
            "Pitcher performance will appear when saved player "
            "results include pitcher roster positions."
        )
    else:
        pitcher_summary = (
            pitchers.groupby("player", as_index=False)
            .agg(
                uses=("lineup_key", "nunique"),
                entries=("player", "size"),
                average_fpts=("fpts", "mean"),
                average_ownership=("ownership", "mean"),
                total_fees=("entry_fee", "sum"),
                profit=("profit", "sum"),
                cash_rate=("cashed", "mean"),
            )
        )

        pitcher_summary = pitcher_summary[
            pitcher_summary["player"]
            .astype(str)
            .str.strip()
            != ""
        ].copy()

        pitcher_summary["roi"] = pitcher_summary.apply(
            lambda row: (
                row["profit"] / row["total_fees"]
                if row["total_fees"] > 0
                else 0.0
            ),
            axis=1,
        )

        pitcher_summary = pitcher_summary.sort_values(
            ["uses", "average_fpts"],
            ascending=[False, False],
        )

        pitcher_display = pitcher_summary[
            [
                "player",
                "uses",
                "entries",
                "average_fpts",
                "average_ownership",
                "profit",
                "roi",
                "cash_rate",
            ]
        ].copy()

        pitcher_display["average_fpts"] = (
            pitcher_display["average_fpts"].round(2)
        )

        pitcher_display["average_ownership"] = (
            pitcher_display["average_ownership"].map(
                lambda value: f"{value:.1f}%"
            )
        )

        pitcher_display["profit"] = (
            pitcher_display["profit"].map(format_profit)
        )

        pitcher_display["roi"] = pitcher_display["roi"].map(
            lambda value: f"{value:.1%}"
        )

        pitcher_display["cash_rate"] = (
            pitcher_display["cash_rate"].map(
                lambda value: f"{value:.1%}"
            )
        )

        pitcher_display = pitcher_display.rename(
            columns={
                "player": "Pitcher",
                "uses": "Uses",
                "entries": "Entries",
                "average_fpts": "Avg DK Points",
                "average_ownership": "Avg Ownership",
                "profit": "Profit",
                "roi": "ROI",
                "cash_rate": "Cash Rate",
            }
        )

        st.caption(
            "Uses counts unique lineups. Entries counts every "
            "contest entry containing that pitcher."
        )

        st.dataframe(
            pitcher_display,
            use_container_width=True,
            hide_index=True,
            height=360,
        )


    st.divider()

    st.subheader("📊 Team Stack Performance")

    stack_players = players.copy()

    stack_players["team"] = (
        stack_players["team"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    stack_players["normalized_position"] = (
        stack_players["roster_position"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    stack_players = stack_players[
        ~stack_players["normalized_position"].isin(
            {"P", "SP", "RP"}
        )
    ].copy()

    stack_players = stack_players[
        stack_players["team"] != ""
    ].copy()

    if stack_players.empty:
        st.info(
            "Team Stack Performance will appear when saved "
            "player results include team abbreviations."
        )
    else:
        contest_group_columns = [
            "slate_date",
            "contest_type",
            "entry_name",
            "lineup_id",
            "lineup_key",
            "entry_fee",
            "winnings",
            "profit",
            "cashed",
        ]

        team_lineup_results = (
            stack_players.groupby(
                contest_group_columns + ["team"],
                dropna=False,
                as_index=False,
            )
            .agg(
                hitter_count=("player", "nunique"),
                team_dk_points=("fpts", "sum"),
            )
        )

        team_lineup_results = team_lineup_results.sort_values(
            [
                "lineup_key",
                "contest_type",
                "entry_name",
                "hitter_count",
                "team_dk_points",
                "team",
            ],
            ascending=[True, True, True, False, False, True],
        )

        primary_stacks = (
            team_lineup_results.groupby(
                [
                    "lineup_key",
                    "contest_type",
                    "entry_name",
                ],
                dropna=False,
                as_index=False,
            )
            .first()
        )

        primary_stacks["stack"] = (
            primary_stacks["team"]
            + " "
            + primary_stacks["hitter_count"]
            .astype(int)
            .astype(str)
            + "-man"
        )

        stack_summary = (
            primary_stacks.groupby(
                ["team", "hitter_count", "stack"],
                as_index=False,
            )
            .agg(
                uses=("lineup_key", "nunique"),
                entries=("stack", "size"),
                average_team_fpts=("team_dk_points", "mean"),
                average_lineup_fpts=("lineup_key", "size"),
                total_fees=("entry_fee", "sum"),
                profit=("profit", "sum"),
                cash_rate=("cashed", "mean"),
            )
        )

        lineup_points_by_entry = (
            players.groupby(
                [
                    "lineup_key",
                    "contest_type",
                    "entry_name",
                ],
                dropna=False,
                as_index=False,
            )
            .agg(lineup_dk_points=("fpts", "sum"))
        )

        primary_stacks = primary_stacks.merge(
            lineup_points_by_entry,
            on=[
                "lineup_key",
                "contest_type",
                "entry_name",
            ],
            how="left",
        )

        average_lineup_points = (
            primary_stacks.groupby(
                ["team", "hitter_count", "stack"],
                as_index=False,
            )
            .agg(
                average_lineup_fpts=(
                    "lineup_dk_points",
                    "mean",
                )
            )
        )

        stack_summary = stack_summary.drop(
            columns=["average_lineup_fpts"]
        ).merge(
            average_lineup_points,
            on=["team", "hitter_count", "stack"],
            how="left",
        )

        stack_summary["roi"] = stack_summary.apply(
            lambda row: (
                row["profit"] / row["total_fees"]
                if row["total_fees"] > 0
                else 0.0
            ),
            axis=1,
        )

        stack_summary = stack_summary.sort_values(
            ["uses", "average_team_fpts"],
            ascending=[False, False],
        )

        stack_display = stack_summary[
            [
                "stack",
                "uses",
                "entries",
                "average_team_fpts",
                "average_lineup_fpts",
                "profit",
                "roi",
                "cash_rate",
            ]
        ].copy()

        stack_display["average_team_fpts"] = (
            stack_display["average_team_fpts"].round(2)
        )

        stack_display["average_lineup_fpts"] = (
            stack_display["average_lineup_fpts"].round(2)
        )

        stack_display["profit"] = stack_display["profit"].map(
            format_profit
        )

        stack_display["roi"] = stack_display["roi"].map(
            lambda value: f"{value:.1%}"
        )

        stack_display["cash_rate"] = (
            stack_display["cash_rate"].map(
                lambda value: f"{value:.1%}"
            )
        )

        stack_display = stack_display.rename(
            columns={
                "stack": "Primary Stack",
                "uses": "Uses",
                "entries": "Entries",
                "average_team_fpts": "Avg Stack DK Points",
                "average_lineup_fpts": "Avg Lineup DK Points",
                "profit": "Profit",
                "roi": "ROI",
                "cash_rate": "Cash Rate",
            }
        )

        st.caption(
            "The primary stack is the team with the most hitters "
            "in each lineup. Ties are broken by team DK points."
        )

        st.dataframe(
            stack_display,
            use_container_width=True,
            hide_index=True,
            height=380,
        )